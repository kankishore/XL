# Read data from XL & Generate in JSON Format
import json
import openpyxl

def generate_json(row,):
    xl = openpyxl.load_workbook("XL_Driver.xlsx")
    sheet = xl["SmokeTest"]

    data = json.dumps({
    "properties": [
    {
      "property": "email",
      "value": sheet.cell(row = row, column = 7).value
    },
    {
      "property": "firstname",
      "value": sheet.cell(row = row, column = 8).value
    },
      {
          "property": "lastname",
          "value": sheet.cell(row = row, column = 9).value
      },
      {
          "property": "company",
          "value": sheet.cell(row = row, column = 10).value
      }
    ]})
    data_json=json.loads(data)
    print(data_json)
    print(type(data_json))
    print(data_json['properties'][1])
    return data