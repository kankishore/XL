
import openpyxl

driverXL=openpyxl.load_workbook("XL_Driver.xlsx")
sheet=driverXL['SmokeTest']
cell = sheet.cell(row = 1, column = 1)
print(cell.value)

# get no. of rows
print(sheet.max_row)
# get no. of columns
print(sheet.max_column)

