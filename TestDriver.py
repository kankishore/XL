import openpyxl
import requests
import configparser
import json
from BaseClass import BaseClass
from TestEnvironment import TestEnvironment
import allure
import pytest
import unittest
import JSON_Helper

class TestDriver(BaseClass,TestEnvironment,unittest.TestCase):
    allure.testcase("Driver")
    def test_execute(self):
        for i in range(2,self.sheet.max_row+1):
            with allure.step(self.sheet.cell(i,self.TCID).value):
                cell=self.sheet.cell(row=i,column=self.URL)
                REQUEST_URL=self.BASE_URL+str(cell.value)+self.API_KEY
                headers = {}
                headers["Content-Type"] = "application/json"

                if(self.sheet.cell(i,self.METHOD).value=='GET'):
                    res=requests.get(REQUEST_URL)
                    print(self.sheet.cell(i,self.EXPRC).value)
                    assert res.status_code==self.sheet.cell(i,self.EXPRC).value

                elif(self.sheet.cell(i,self.METHOD).value=='POST'):
                    data=JSON_Helper.generate_json(i)
                    #print(data)
                    res=requests.post(REQUEST_URL,data, headers)
                    print(res.text)

            try:
                assert res.status_code==self.sheet.cell(i,self.EXPRC).value
                print("Test PASSED - ",self.sheet.cell(i,self.TCID).value)
            except AssertionError as e:
                pass
                #print("Test FAILED due to Assertion Error - ",self.sheet.cell(i,self.TCID).value)
                #print("Exp Value: ",self.sheet.cell(i,self.EXPRC).value, " Actual Value: ",res.status_code)

#t=TestDriver()
#t.test_execute()